# -*- coding: utf-8 -*-
"""
Excel / CSV 輸出模組：使用 openpyxl 取代 xlwt，輸出 .xlsx 格式
原始：HETA.py:424-516
"""

import csv
import networkx as nx
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment

from heta.constants import *


# 樣式定義（對應原始 xlwt 的 si, st, sb）
FONT_HEADER = Font(name='Arial', color='8B0000', bold=True)    # dark_red bold
FONT_BODY = Font(name='Arial', color='00008B')                  # dark_blue
ALIGN_LEFT = Alignment(horizontal='left')
ALIGN_CENTER = Alignment(horizontal='center')


def _write_cell(sheet, row, col, value, font=None, alignment=None):
    """
    寫入儲存格（1-based 索引，與 openpyxl 一致）
    注意：原始 xlwt 使用 0-based，這裡的 row/col 參數也用 0-based 傳入，
    內部轉換為 1-based。
    """
    cell = sheet.cell(row=row + 1, column=col + 1, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment


def write_link_analysis_excel(result, output_path):
    """
    將連結分析結果寫入 Excel 檔案

    參數：
        result: LinkAnalysisResult 物件
        output_path: 輸出 .xlsx 檔案路徑
    """
    g = result.graph
    layers = result.layers

    wb = openpyxl.Workbook()
    # 移除預設 sheet
    wb.remove(wb.active)

    s1 = wb.create_sheet('target network')
    s2 = wb.create_sheet(f'{len(result.random_network_stats)} random networks')
    s3 = wb.create_sheet('node information')

    # Phase 1.1: 標題資料
    row = 5
    col = 7

    _write_cell(s1, 0, 0, f'number of nodes = {result.num_nodes}', FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 1, 0, f'number of edges = {result.num_edges}', FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 2, 0, f'average degree = {result.avg_degree}', FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 3, 0, f'diameter = {result.diameter}', FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 4, 0, f'average shortest path = {result.avg_shortest_path}', FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 5, 0, f'average clustering coefficient = {result.avg_clustering_coeff}', FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 6, 0, f'degree assortativity coefficient = {result.degree_assortativity}', FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 7, 0,
                f'BOND = {result.bond_count} ({100 * round(float(result.bond_count) / result.num_edges, 4)}%)',
                FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 8, 0,
                f'sink = {result.sink_count} ({100 * round(float(result.sink_count) / result.num_edges, 4)}%)',
                FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 9, 0,
                f'local bridge = {result.local_bridge_count} ({100 * round(float(result.local_bridge_count) / result.num_edges, 4)}%)',
                FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 10, 0,
                f'global bridge = {result.global_bridge_count} ({100 * round(float(result.global_bridge_count) / result.num_edges, 4)}%)',
                FONT_HEADER, ALIGN_LEFT)
    _write_cell(s1, 11, 0, f'entropy = {result.graph_entropy}', FONT_HEADER, ALIGN_LEFT)

    _write_cell(s1, row - 1, col - 6, 'st.sp', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s1, row - 1, col - 5, 'avg.sp', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s1, row - 1, col - 4, 's.cc', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s1, row - 1, col - 3, 't.cc', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s1, row - 1, col - 2, 'source', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s1, row - 1, col - 1, 'target', FONT_HEADER, ALIGN_CENTER)

    # 最終分類欄位（置於所有 layer 欄之後）
    final_col = col + (layers * 2)
    _write_cell(s1, row - 1, final_col, 'final type', FONT_HEADER, ALIGN_CENTER)

    for i in range(layers):
        r1_val = round(result.thresholds_r1.get(str(i + 1), 0), 4)
        r2_val = round(result.thresholds_r2.get(str(i + 1), 0), 4)
        _write_cell(s1, row - 3, col + (i * 2), f'R1 = {r1_val}', FONT_HEADER, ALIGN_LEFT)
        _write_cell(s1, row - 2, col + (i * 2), f'R2 = {r2_val}', FONT_HEADER, ALIGN_LEFT)
        _write_cell(s1, row - 1, col + (i * 2), 'intersection weight', FONT_HEADER, ALIGN_CENTER)
        _write_cell(s1, row - 1, col + (i * 2) + 1, f'layer {i + 1}', FONT_HEADER, ALIGN_CENTER)

    # Phase 1.2: 邊資料
    # 預先計算最短路徑長度（供 st.sp 欄位使用）
    all_shortest_paths = dict(nx.all_pairs_shortest_path_length(g))
    avg_sp = result.avg_shortest_path
    last_layer_key = EDGE_KEY_LAYER + str(layers)

    for s, t in g.edges():
        st_sp = all_shortest_paths.get(s, {}).get(t, 1)
        _write_cell(s1, row, col - 6, st_sp, FONT_BODY)
        _write_cell(s1, row, col - 5, avg_sp, FONT_BODY)
        _write_cell(s1, row, col - 4, round(nx.clustering(g, s), 2), FONT_BODY)
        _write_cell(s1, row, col - 3, round(nx.clustering(g, t), 2), FONT_BODY)
        _write_cell(s1, row, col - 2, s, FONT_BODY)
        _write_cell(s1, row, col - 1, t, FONT_BODY)

        for i in range(layers):
            _write_cell(s1, row, col + (i * 2), round(g[s][t][-(i + 1)], 4), FONT_BODY)
            if i == 0:
                _write_cell(s1, row, col + (i * 2) + 1,
                            g[s][t][EDGE_KEY_LAYER + str(i + 1)], FONT_BODY)
            elif g[s][t][EDGE_KEY_LAYER + str(i + 1)] != g[s][t][EDGE_KEY_LAYER + str(i)]:
                _write_cell(s1, row, col + (i * 2) + 1,
                            g[s][t][EDGE_KEY_LAYER + str(i + 1)], FONT_BODY)
            else:
                _write_cell(s1, row, col + (i * 2) + 1, '...', FONT_BODY)

        # 最終分類（最後一層的判定結果）
        _write_cell(s1, row, final_col, g[s][t][last_layer_key], FONT_HEADER)
        row += 1

    # Phase 2: 隨機網絡統計資料
    rn_stats = result.random_network_stats
    actual_times = len(rn_stats)
    row = 5
    col = 3

    for i in range(layers):
        l = str(i + 1)
        _write_cell(s2, row - 2, col + (i * 4), f'layer {l}', FONT_HEADER, ALIGN_CENTER)
        _write_cell(s2, row - 1, col + (i * 4), 'AVG', FONT_HEADER, ALIGN_CENTER)
        _write_cell(s2, row - 1, col + (i * 4) + 1, 'STD', FONT_HEADER, ALIGN_CENTER)
        for j in range(actual_times):
            _write_cell(s2, row + j, col + (i * 4),
                        rn_stats[j].get(GRAPH_KEY_AVG_COMMON_NODES + l, 0), FONT_BODY)
            _write_cell(s2, row + j, col + (i * 4) + 1,
                        rn_stats[j].get(GRAPH_KEY_STD_COMMON_NODES + l, 0), FONT_BODY)

    # Phase 3: 節點資訊熵
    row = 1
    col = 1
    now = 1

    _write_cell(s3, row, col + 0, 'node', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row, col + 1, 'degree', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row - 1, col + 2, 'o.entropy = ', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row, col + 2, 'n.entropy', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row - 1, col + 3, result.graph_entropy, FONT_BODY)
    _write_cell(s3, row, col + 3, 'gain', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row - 1, col + 4, g.graph[GRAPH_KEY_EDGE_CLASS][BOND], FONT_BODY)
    _write_cell(s3, row, col + 4, 'BOND', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row - 1, col + 5, g.graph[GRAPH_KEY_EDGE_CLASS][LOCAL_BRIDGE], FONT_BODY)
    _write_cell(s3, row, col + 5, 'local bridge', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row - 1, col + 6, g.graph[GRAPH_KEY_EDGE_CLASS][GLOBAL_BRIDGE], FONT_BODY)
    _write_cell(s3, row, col + 6, 'global bridge', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row - 1, col + 7, g.graph[GRAPH_KEY_EDGE_CLASS].get(SINK, 0), FONT_BODY)
    _write_cell(s3, row, col + 7, 'sink', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row, col + 8, 'avg', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s3, row + 1, col + 8, result.node_info_avg, FONT_BODY)

    for s in g.nodes():
        _write_cell(s3, row + now, col + 0, s, FONT_BODY)
        _write_cell(s3, row + now, col + 1, g.degree(s), FONT_BODY)
        _write_cell(s3, row + now, col + 2, g.nodes[s][NODE_KEY_NEW_ENTROPY], FONT_BODY)
        _write_cell(s3, row + now, col + 3, g.nodes[s][NODE_KEY_INFORMATION_GAIN], FONT_BODY)
        _write_cell(s3, row + now, col + 4, g.nodes[s][NODE_KEY_EDGE_CLASS][BOND], FONT_BODY)
        _write_cell(s3, row + now, col + 5, g.nodes[s][NODE_KEY_EDGE_CLASS][LOCAL_BRIDGE], FONT_BODY)
        _write_cell(s3, row + now, col + 6, g.nodes[s][NODE_KEY_EDGE_CLASS][GLOBAL_BRIDGE], FONT_BODY)
        _write_cell(s3, row + now, col + 7, g.nodes[s][NODE_KEY_EDGE_CLASS].get(SINK, 0), FONT_BODY)
        now += 1

    wb.save(output_path)


def write_suite_experiment_excel(suite_result, output_path):
    """
    將批次實驗結果寫入 Excel 檔案

    參數：
        suite_result: SuiteExperimentResult 物件
        output_path: 輸出 .xlsx 檔案路徑
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # D2: 按照階層聚類順序排列指紋資料（與 Fingerprint 圖表一致）
    if suite_result.corr_index and len(suite_result.corr_index) == len(suite_result.labels):
        order = suite_result.corr_index
        sorted_labels = [suite_result.labels[i] for i in order]
        sorted_bar = {
            k: v[order] if hasattr(v, '__getitem__') and hasattr(v, 'dtype') else v
            for k, v in suite_result.bar_data.items()
        }
    else:
        sorted_labels = suite_result.labels
        sorted_bar = suite_result.bar_data

    # Sheet 1: 指紋資料（按聚類順序）
    s1 = wb.create_sheet('fingerprints')
    _write_cell(s1, 0, 0, 'Network', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s1, 0, 1, 'BOND', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s1, 0, 2, 'Local Bridge', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s1, 0, 3, 'Global Bridge', FONT_HEADER, ALIGN_CENTER)
    _write_cell(s1, 0, 4, 'Sink', FONT_HEADER, ALIGN_CENTER)

    for i, label in enumerate(sorted_labels):
        _write_cell(s1, i + 1, 0, label, FONT_BODY)
        _write_cell(s1, i + 1, 1, float(sorted_bar[BOND][i]), FONT_BODY)
        _write_cell(s1, i + 1, 2, float(sorted_bar[LOCAL_BRIDGE][i]), FONT_BODY)
        _write_cell(s1, i + 1, 3, float(sorted_bar[GLOBAL_BRIDGE][i]), FONT_BODY)
        _write_cell(s1, i + 1, 4, float(sorted_bar[SINK][i]), FONT_BODY)

    # Sheet 2: 相關矩陣
    s2 = wb.create_sheet('correlation matrix')
    _write_cell(s2, 0, 0, '', FONT_HEADER)
    for i, label in enumerate(suite_result.corr_labels):
        _write_cell(s2, 0, i + 1, label, FONT_HEADER, ALIGN_CENTER)
        _write_cell(s2, i + 1, 0, label, FONT_HEADER)
    for i in range(len(suite_result.corr_labels)):
        for j in range(len(suite_result.corr_labels)):
            _write_cell(s2, i + 1, j + 1, round(float(suite_result.corr_matrix[i, j]), 4), FONT_BODY)

    # D1: Sheet 3: 網絡基本統計量（按聚類順序）
    s3 = wb.create_sheet('network statistics')
    stat_headers = [
        ('Network', 'network'),
        ('Nodes', 'nodes'),
        ('Edges', 'edges'),
        ('Avg Degree', 'avg_degree'),
        ('Diameter', 'diameter'),
        ('Avg Shortest Path', 'avg_shortest_path'),
        ('Avg Clustering Coeff', 'avg_clustering_coeff'),
        ('Degree Assortativity', 'degree_assortativity'),
        ('Entropy', 'entropy'),
    ]
    for c, (header, _) in enumerate(stat_headers):
        _write_cell(s3, 0, c, header, FONT_HEADER, ALIGN_CENTER)

    net_stats = suite_result.network_stats
    for i, label in enumerate(sorted_labels):
        _write_cell(s3, i + 1, 0, label, FONT_BODY)
        # 搜尋對應的 stats key（label + '_1' 或含路徑的 key）
        stats_key = label + '_1'
        stats = net_stats.get(stats_key, {})
        if not stats:
            for k in net_stats:
                if k.endswith(stats_key):
                    stats = net_stats[k]
                    break
        for c, (_, field) in enumerate(stat_headers):
            if c == 0:
                continue
            val = stats.get(field, '')
            if val != '':
                _write_cell(s3, i + 1, c, val, FONT_BODY)

    wb.save(output_path)


def write_edge_classification_csv(result, output_path):
    """
    將邊分類結果匯出為 CSV 檔案，可供 Gephi / Cytoscape 等工具匯入。

    參數：
        result: LinkAnalysisResult 物件
        output_path: 輸出 .csv 檔案路徑

    輸出格式：
        source, target, type, weight
    """
    g = result.graph
    layers = result.layers
    last_layer_key = EDGE_KEY_LAYER + str(layers)

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['source', 'target', 'type', 'weight'])
        for s, t in g.edges():
            edge_type = g[s][t].get(last_layer_key, '')
            # 取得最後一層的 intersection weight 作為權重
            weight = round(g[s][t].get(-(layers), 0), 4)
            writer.writerow([s, t, edge_type, weight])
